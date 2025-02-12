"""
    This file generates a list of translations from en-xx based on the 'train' partition of each XLSX file in the specified directory.
    For every row in the 'train' partition, it extracts the 'id' and 'utterance' columns and appends them to a results list.
"""

import os
import json
from openpyxl import load_workbook
from zipfile import BadZipFile
import logging
from absl import app
from flags_config import FLAGS

logging.basicConfig(level=logging.INFO)

def generate_translations_from_en_xx(processed_files_dir):
    """
    Generates all translations from en-xx for all train datasets.
    """
    files = os.listdir(processed_files_dir)
    results = []

    logging.info("Processing Files")

    for file in files:
        if not file.endswith('.xlsx'):
            continue

        file_path = os.path.join(processed_files_dir, file)

        try:
            wb = load_workbook(file_path, read_only=True)
            sheet = wb.active

            # Assuming columns: id, utterance, partition (in any order)
            headers = [cell.value for cell in sheet[1]]

            id_col_idx = headers.index('id')
            utterance_col_idx = headers.index('utterance')
            partition_col_idx = headers.index('partition')

            for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip header
                if row[partition_col_idx] == "train":
                    results.append({"id": row[id_col_idx], "utterance": row[utterance_col_idx]})
        except BadZipFile:
            continue

    return results

# Usage:
# processed_files_dir = FLAGS.processed_files_dir
translations = generate_translations_from_en_xx(FLAGS.processed_files_dir)

# Save to all_translations.json file without using Pandas
with open(FLAGS.english_dataset, 'w', encoding='utf-8') as f:
    json.dump(translations, f, ensure_ascii=False, indent=4)

logging.info('Translations have been saved to all_translations.json')
